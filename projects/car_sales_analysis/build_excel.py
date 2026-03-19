"""
Rebuild Car Sales Analysis Excel workbook — all 30 manufacturers, correct titles.
"""
import csv, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Load CSV ──────────────────────────────────────────────────
script_dir = os.path.dirname(os.path.abspath(__file__))
csv_path   = os.path.join(script_dir, 'Car_sales_row_data.csv')

with open(csv_path) as f:
    raw = list(csv.DictReader(f))

def sf(v):
    try: return float(v) if v.strip() else None
    except: return None

rows = []
for r in raw:
    price  = sf(r['Price_in_thousands'])
    resale = sf(r['__year_resale_value'])
    sales  = sf(r['Sales_in_thousands'])
    rows.append({
        'manufacturer': r['Manufacturer'],
        'model':        r['Model'],
        'price':        price,
        'resale':       resale,
        'sales':        sales,
        'hp':           sf(r['Horsepower']),
        'mpg':          sf(r['Fuel_efficiency']),
        'ppf':          sf(r['Power_perf_factor']),
        'retention':    round(resale / price * 100, 2) if price and resale else None,
        'depreciation': round(price - resale, 2)       if price and resale else None,
    })

# ── Manufacturer summary ──────────────────────────────────────
mfr_map = defaultdict(list)
for r in rows:
    if r['retention'] is not None:
        mfr_map[r['manufacturer']].append(r)

mfr_summary = []
for mfr, models in mfr_map.items():
    mfr_summary.append({
        'manufacturer':    mfr,
        'models':          len(models),
        'avg_price':       round(sum(m['price']  for m in models) / len(models), 2),
        'avg_resale':      round(sum(m['resale'] for m in models) / len(models), 2),
        'avg_retention':   round(sum(m['retention'] for m in models) / len(models), 1),
        'avg_depreciation':round(sum(m['depreciation'] for m in models) / len(models), 2),
        'total_sales':     round(sum(m['sales'] for m in models if m['sales']), 1),
    })
mfr_summary.sort(key=lambda x: x['avg_retention'], reverse=True)

# ── Styles ────────────────────────────────────────────────────
DARK_BLUE   = '1F3864'
MID_BLUE    = '2E75B6'
LIGHT_BLUE  = 'D6E4F0'
WHITE       = 'FFFFFF'
GREEN       = '27AE60'
RED         = 'E74C3C'
YELLOW      = 'F39C12'
GREY        = 'F2F2F2'

def hdr_fill(hex_color): return PatternFill('solid', fgColor=hex_color)
def font(bold=False, size=11, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, col_start, col_end, bg=DARK_BLUE, fg=WHITE, size=11):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = hdr_fill(bg)
        cell.font      = font(bold=True, size=size, color=fg)
        cell.alignment = center()
        cell.border    = border()

def style_data_row(ws, row, col_start, col_end, bg=WHITE):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = hdr_fill(bg)
        cell.font      = font(size=10)
        cell.alignment = center()
        cell.border    = border()

wb = Workbook()

# ════════════════════════════════════════════════════════════════
# SHEET 1 — Raw Data
# ════════════════════════════════════════════════════════════════
ws_raw = wb.active
ws_raw.title = 'Raw Data'
ws_raw.sheet_view.showGridLines = False

headers = ['Manufacturer','Model','Price ($K)','1-Yr Resale ($K)',
           'Retention (%)','Depreciation ($K)','Sales (K units)',
           'Horsepower','MPG','Power Perf Factor']
for c, h in enumerate(headers, 1):
    ws_raw.cell(row=1, column=c, value=h)
style_header_row(ws_raw, 1, 1, len(headers))

for i, r in enumerate(rows, 2):
    data = [r['manufacturer'], r['model'], r['price'], r['resale'],
            r['retention'], r['depreciation'], r['sales'],
            r['hp'], r['mpg'], r['ppf']]
    bg = GREY if i % 2 == 0 else WHITE
    for c, val in enumerate(data, 1):
        ws_raw.cell(row=i, column=c, value=val)
    style_data_row(ws_raw, i, 1, len(headers), bg)
    # Color retention cell
    ret_cell = ws_raw.cell(row=i, column=5)
    if r['retention'] is not None:
        if r['retention'] >= 75:
            ret_cell.fill = hdr_fill('D5F5E3')
        elif r['retention'] >= 65:
            ret_cell.fill = hdr_fill('FDEBD0')
        else:
            ret_cell.fill = hdr_fill('FADBD8')

col_widths = [14, 20, 13, 16, 13, 17, 14, 12, 8, 17]
for c, w in enumerate(col_widths, 1):
    ws_raw.column_dimensions[get_column_letter(c)].width = w
ws_raw.row_dimensions[1].height = 30

# ════════════════════════════════════════════════════════════════
# SHEET 2 — Manufacturer Summary (all 30 brands)
# ════════════════════════════════════════════════════════════════
ws_mfr = wb.create_sheet('Manufacturer Summary')
ws_mfr.sheet_view.showGridLines = False

# Title
ws_mfr.merge_cells('A1:G1')
title_cell = ws_mfr['A1']
title_cell.value     = 'Manufacturer Value Retention Summary — All Brands'
title_cell.font      = font(bold=True, size=14, color=WHITE)
title_cell.fill      = hdr_fill(DARK_BLUE)
title_cell.alignment = center()
ws_mfr.row_dimensions[1].height = 36

# Sub-header note
ws_mfr.merge_cells('A2:G2')
note = ws_mfr['A2']
note.value     = 'Retention Rate = 1-Year Resale Value ÷ Original Price × 100  |  Sorted by retention (high → low)'
note.font      = font(italic=True, size=9, color='555555')
note.alignment = center()
note.fill      = hdr_fill(LIGHT_BLUE)
ws_mfr.row_dimensions[2].height = 18

headers2 = ['Manufacturer','# Models','Avg Price ($K)','Avg Resale ($K)',
            'Avg Retention (%)','Avg Depreciation ($K)','Total Sales (K)']
for c, h in enumerate(headers2, 1):
    ws_mfr.cell(row=3, column=c, value=h)
style_header_row(ws_mfr, 3, 1, len(headers2), bg=MID_BLUE)
ws_mfr.row_dimensions[3].height = 28

for i, m in enumerate(mfr_summary, 4):
    bg = GREY if i % 2 == 0 else WHITE
    data = [m['manufacturer'], m['models'], m['avg_price'], m['avg_resale'],
            m['avg_retention'], m['avg_depreciation'], m['total_sales']]
    for c, val in enumerate(data, 1):
        cell = ws_mfr.cell(row=i, column=c, value=val)
        cell.fill      = hdr_fill(bg)
        cell.font      = font(size=10, bold=(c==1))
        cell.alignment = left() if c == 1 else center()
        cell.border    = border()
    # Color retention cell
    ret = m['avg_retention']
    ret_cell = ws_mfr.cell(row=i, column=5)
    if ret >= 75:
        ret_cell.fill = hdr_fill('D5F5E3')
        ret_cell.font = font(bold=True, size=10, color='1E8449')
    elif ret >= 65:
        ret_cell.fill = hdr_fill('FDEBD0')
        ret_cell.font = font(bold=True, size=10, color='935116')
    else:
        ret_cell.fill = hdr_fill('FADBD8')
        ret_cell.font = font(bold=True, size=10, color='922B21')
    ws_mfr.row_dimensions[i].height = 20

col_widths2 = [18, 10, 14, 14, 16, 18, 14]
for c, w in enumerate(col_widths2, 1):
    ws_mfr.column_dimensions[get_column_letter(c)].width = w

# ── Bar Chart: Retention by Manufacturer ──────────────────────
chart = BarChart()
chart.type    = 'bar'
chart.title   = 'Average 1-Year Value Retention by Manufacturer (%)'
chart.y_axis.title = 'Manufacturer'
chart.x_axis.title = 'Retention Rate (%)'
chart.style   = 10
chart.width   = 22
chart.height  = 18

n = len(mfr_summary)
data_ref  = Reference(ws_mfr, min_col=5, min_row=3, max_row=3+n)
cats_ref  = Reference(ws_mfr, min_col=1, min_row=4, max_row=3+n)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill   = MID_BLUE
chart.series[0].graphicalProperties.line.solidFill = MID_BLUE

ws_mfr.add_chart(chart, 'I3')

# ════════════════════════════════════════════════════════════════
# SHEET 3 — Top & Bottom Models
# ════════════════════════════════════════════════════════════════
ws_models = wb.create_sheet('Top & Bottom Models')
ws_models.sheet_view.showGridLines = False

complete_rows = [r for r in rows if r['retention'] is not None and r['sales']]
top10    = sorted(complete_rows, key=lambda x: x['retention'], reverse=True)[:10]
bottom10 = sorted(complete_rows, key=lambda x: x['retention'])[:10]

# Title
ws_models.merge_cells('A1:F1')
t = ws_models['A1']
t.value     = 'Model-Level Analysis: Best & Worst Value Retention'
t.font      = font(bold=True, size=14, color=WHITE)
t.fill      = hdr_fill(DARK_BLUE)
t.alignment = center()
ws_models.row_dimensions[1].height = 36

# TOP 10
ws_models.merge_cells('A2:F2')
ws_models['A2'].value     = '🏆 Top 10 Models — Best Value Retention'
ws_models['A2'].font      = font(bold=True, size=11, color=WHITE)
ws_models['A2'].fill      = hdr_fill(GREEN)
ws_models['A2'].alignment = center()

top_headers = ['Manufacturer','Model','Price ($K)','Resale ($K)','Retention (%)','Sales (K)']
for c, h in enumerate(top_headers, 1):
    ws_models.cell(row=3, column=c, value=h)
style_header_row(ws_models, 3, 1, 6, bg='1E8449', fg=WHITE)

for i, r in enumerate(top10, 4):
    bg = 'EAFAF1' if i % 2 == 0 else WHITE
    data = [r['manufacturer'], r['model'], r['price'], r['resale'], r['retention'], r['sales']]
    for c, val in enumerate(data, 1):
        cell = ws_models.cell(row=i, column=c, value=val)
        cell.fill      = hdr_fill(bg)
        cell.font      = font(size=10)
        cell.alignment = center()
        cell.border    = border()
    ws_models.row_dimensions[i].height = 18

# BOTTOM 10
start = 15
ws_models.merge_cells(f'A{start}:F{start}')
ws_models[f'A{start}'].value     = '⚠️ Bottom 10 Models — Highest Depreciation Risk'
ws_models[f'A{start}'].font      = font(bold=True, size=11, color=WHITE)
ws_models[f'A{start}'].fill      = hdr_fill(RED)
ws_models[f'A{start}'].alignment = center()

for c, h in enumerate(top_headers, 1):
    ws_models.cell(row=start+1, column=c, value=h)
style_header_row(ws_models, start+1, 1, 6, bg='922B21', fg=WHITE)

for i, r in enumerate(bottom10, start+2):
    bg = 'FDEDEC' if i % 2 == 0 else WHITE
    data = [r['manufacturer'], r['model'], r['price'], r['resale'], r['retention'], r['sales']]
    for c, val in enumerate(data, 1):
        cell = ws_models.cell(row=i, column=c, value=val)
        cell.fill      = hdr_fill(bg)
        cell.font      = font(size=10)
        cell.alignment = center()
        cell.border    = border()
    ws_models.row_dimensions[i].height = 18

col_widths3 = [14, 22, 11, 12, 13, 10]
for c, w in enumerate(col_widths3, 1):
    ws_models.column_dimensions[get_column_letter(c)].width = w

# ── Chart: Top 10 retention ───────────────────────────────────
chart2 = BarChart()
chart2.type   = 'bar'
chart2.title  = 'Top 10 Models by Value Retention (%)'
chart2.style  = 10
chart2.width  = 18
chart2.height = 12
data2  = Reference(ws_models, min_col=5, min_row=3, max_row=13)
cats2  = Reference(ws_models, min_col=2, min_row=4, max_row=13)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = '1E8449'
ws_models.add_chart(chart2, 'H3')

# ════════════════════════════════════════════════════════════════
# SHEET 4 — Market Positioning
# ════════════════════════════════════════════════════════════════
ws_pos = wb.create_sheet('Market Positioning')
ws_pos.sheet_view.showGridLines = False

ws_pos.merge_cells('A1:G1')
t2 = ws_pos['A1']
t2.value     = 'Market Positioning: Price vs. Sales Volume vs. Retention'
t2.font      = font(bold=True, size=14, color=WHITE)
t2.fill      = hdr_fill(DARK_BLUE)
t2.alignment = center()
ws_pos.row_dimensions[1].height = 36

pos_headers = ['Manufacturer','Model','Price ($K)','Sales (K units)',
               'Retention (%)','Depreciation ($K)','Segment']
for c, h in enumerate(pos_headers, 1):
    ws_pos.cell(row=2, column=c, value=h)
style_header_row(ws_pos, 2, 1, 7, bg=MID_BLUE)

def segment(price, sales):
    med_p, med_s = 22, 50
    if price >= med_p and sales >= med_s: return 'High-Volume Premium'
    if price < med_p  and sales >= med_s: return 'High-Volume Budget'
    if price >= med_p and sales < med_s:  return 'Niche Premium'
    return 'Niche Budget'

seg_colors = {
    'High-Volume Premium': 'D6EAF8',
    'High-Volume Budget':  'D5F5E3',
    'Niche Premium':       'FDEBD0',
    'Niche Budget':        'FDEDEC',
}

pos_rows = [r for r in rows if r['price'] and r['sales'] and r['retention']]
pos_rows.sort(key=lambda x: x['sales'], reverse=True)

for i, r in enumerate(pos_rows, 3):
    seg = segment(r['price'], r['sales'])
    data = [r['manufacturer'], r['model'], r['price'], r['sales'],
            r['retention'], r['depreciation'], seg]
    bg = seg_colors.get(seg, WHITE)
    for c, val in enumerate(data, 1):
        cell = ws_pos.cell(row=i, column=c, value=val)
        cell.fill      = hdr_fill(bg)
        cell.font      = font(size=10)
        cell.alignment = center()
        cell.border    = border()
    ws_pos.row_dimensions[i].height = 18

col_widths4 = [14, 22, 11, 14, 13, 16, 20]
for c, w in enumerate(col_widths4, 1):
    ws_pos.column_dimensions[get_column_letter(c)].width = w

# ════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════
out_path = os.path.join(script_dir, 'Car_sales_analysis.xlsx')
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Sheets: {[s.title for s in wb.worksheets]}')
